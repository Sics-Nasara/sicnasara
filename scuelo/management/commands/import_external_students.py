from __future__ import unicode_literals

from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from pathlib import Path
from django.utils.dateparse import parse_date
from django.utils import timezone
from scuelo.models import AnneeScolaire, Eleve, Inscription, Classe, Mouvement, Ecole, TypeClasse
import datetime
class Command(BaseCommand):
    help = '''Imports data from "External students.xlsx" file.'''

    def handle(self, *args, **options):
        BASE_DIR = str(Path(__file__).resolve().parent.parent.parent)
        full_path = f'{BASE_DIR}/export_sics/External students.xlsx'
        print('Import SICS START %s' % full_path)
        wb = load_workbook(full_path)
        
        # Load the current and previous school years
        annee_scolaire_actuel = AnneeScolaire.objects.get(actuel=True)
        derniere_annee_scolaire = AnneeScolaire.objects.get(actuel=False)
        
        # Import data from Scuole sheet
        ws_scuole = wb['Scuole']
        for row in ws_scuole.iter_rows(min_row=2, values_only=True):
            try:
                new_ecole, created = Ecole.objects.get_or_create(
                    nom=row[1].strip(),  # Strip leading/trailing whitespace
                    defaults={
                        'ville': 'Unknown',  # Update with actual data if available
                        'nom_du_referent': 'Unknown',  # Update with actual data if available
                        'prenom_du_referent': 'Unknown',  # Update with actual data if available
                        'email_du_referent': 'unknown@example.com',  # Update with actual data if available
                        'telephone_du_referent': '0000000000',  # Update with actual data if available
                        'note': row[3] if row[3] else '',
                        'externe': True
                    }
                )
                print(f'Ecole: {new_ecole.nom} created' if created else f'Ecole: {new_ecole.nom} already exists')
            except Exception as ex:
                print(f'Error creating Ecole: {str(ex)}')
        
        # Manually register classes
        classe_data = [
            {"legacy_id": "_PK-6me-WP", "ecole_nom": "Ecole WEND LA MANEGDA", "type_classe_nom": "6me"},
            {"legacy_id": "_PK-6me-LMS", "ecole_nom": "Lycee Municipal de saaba", "type_classe_nom": "6me"},
            {"legacy_id": "_PK-6me-WLM", "ecole_nom": "Ecole WEND LA MANEGDA", "type_classe_nom": "6me"},
            {"legacy_id": "_PK-5me-WP", "ecole_nom": "Ecole WEND LA MANEGDA", "type_classe_nom": "5me"},
            {"legacy_id": "_PK-5me-WLM", "ecole_nom": "Ecole WEND LA MANEGDA", "type_classe_nom": "5me"},
            {"legacy_id": "_PK-5me-LMS", "ecole_nom": "Lycee Municipal de saaba", "type_classe_nom": "5me"},
            {"legacy_id": "_PK-4me-WLM", "ecole_nom": "Ecole WEND LA MANEGDA", "type_classe_nom": "4me"},
            {"legacy_id": "_PK-4me-LMS", "ecole_nom": "Lycee Municipal de saaba", "type_classe_nom": "4me"},
            {"legacy_id": "_PK-3me-WLM", "ecole_nom": "Ecole WEND LA MANEGDA", "type_classe_nom": "3me"},
            {"legacy_id": "_PK-Term-LMS", "ecole_nom": "Lycee Municipal de saaba", "type_classe_nom": "Term"},
            {"legacy_id": "_PK-4me-LPMSBD", "ecole_nom": "Lycee Provincial Molla Sanou de Bobo Dioulasso", "type_classe_nom": "4me"},
            {"legacy_id": "_PK-CE2-EPCELP", "ecole_nom": "Ecole Primaire Catholique Effata Ludovic Pavoni", "type_classe_nom": "CE2"},
            {"legacy_id": "_PK-CE2-EPPLE", "ecole_nom": "Ecole Primaire Les Emerites", "type_classe_nom": "CE2"},
            {"legacy_id": "_PK-CM1-LSB", "ecole_nom": "Lycee Municipal de saaba", "type_classe_nom": "CM1"},
            {"legacy_id": "_PK-CM1-YAMT", "ecole_nom": "Ecole Yamtenga", "type_classe_nom": "CM1"},
            # Add any other missing legacy_ids here
        ]
        
        for data in classe_data:
            try:
                ecole = Ecole.objects.get(nom=data['ecole_nom'].strip())
                type_classe, created = TypeClasse.objects.get_or_create(
                    nom=data['type_classe_nom'],
                    defaults={'ordre': 0, 'type_ecole': 'L'}  # Update with actual data if available
                )
                new_classe, created = Classe.objects.get_or_create(
                    legacy_id=data['legacy_id'],
                    defaults={
                        'ecole': ecole,
                        'type': type_classe,
                        'nom': data['type_classe_nom']
                    }
                )
                print(f'Classe: {new_classe.nom} created' if created else f'Classe: {new_classe.nom} already exists')
            except Ecole.DoesNotExist:
                print(f'Error creating Classe: Ecole named {data["ecole_nom"]} does not exist.')
            except Exception as ex:
                print(f'Error creating Classe: {str(ex)}')

        # Function to parse date safely
        def parse_date_safe(date_str):
            if date_str is None:
                return None
            if isinstance(date_str, str):
                return parse_date(date_str)
            if isinstance(date_str, (datetime.date, datetime.datetime)):
                return date_str.date() if isinstance(date_str, datetime.datetime) else date_str
            return None
        
        # Import data from Studente sheet
        ws_studente = wb['Studente']
        for row in ws_studente.iter_rows(min_row=2, values_only=True):
            try:
                classe = Classe.objects.get(legacy_id=row[1])
                new_eleve, created = Eleve.objects.get_or_create(
                    legacy_id=row[0],
                    defaults={
                        'nom': row[5],
                        'prenom': row[4],
                        'date_enquete': parse_date_safe(row[3]),
                        'condition_eleve': row[6][:4],  # Ensure length is not exceeding the limit
                        'sex': row[7],
                        'date_naissance': parse_date_safe(row[8]),
                        'cs_py': row[9],
                        'hand': row[10],
                        'annee_inscr': row[11],
                        'parent': row[12],
                        'tel_parent': row[13],
                        'note_eleve': row[18] if row[18] else '',
                    }
                )
                print(f'Eleve: {new_eleve.nom} {new_eleve.prenom} created' if created else f'Eleve: {new_eleve.nom} {new_eleve.prenom} already exists')
                
                # Create Inscription
                Inscription.objects.get_or_create(
                    eleve=new_eleve,
                    classe=classe,
                    annee_scolaire=annee_scolaire_actuel,
                    defaults={'date_inscription': parse_date_safe(row[19]) or timezone.now()}
                )
                print(f'Inscription for Eleve: {new_eleve.nom} {new_eleve.prenom} created')
            except Classe.DoesNotExist:
                print(f'Error creating Eleve: Classe with legacy_id {row[1]} does not exist.')
            except Exception as ex:
                print(f'Error creating Eleve: {str(ex)}')

        print('Import external students END %s' % full_path)
